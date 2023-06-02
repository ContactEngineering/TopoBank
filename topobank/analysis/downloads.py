"""
Views and helper functions for downloading analyses.
"""

import tempfile

import openpyxl
import pint
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseBadRequest, HttpResponseNotFound
from django.contrib.contenttypes.models import ContentType
from django.utils.text import slugify

import pandas as pd
import io
import numpy as np
import textwrap

from .models import Analysis
from ..manager.models import Surface
from .registry import register_download_function, AnalysisRegistry, UnknownKeyException
from .functions import VIZ_SERIES


#######################################################################
# Download views
#######################################################################


def download_analyses(request, ids, file_format):
    """View returning a file comprised from analyses results.

    Parameters
    ----------
    request: HttpRequest
        The request object
    ids: str
        comma separated string with analyses ids
    file_format: str
        Requested file format, e.g. 'txt' or 'xlsx', depends on what was registerd.

    Returns
    -------
        HttpResponse downloading a file
    """

    #
    # Check permissions and collect analyses
    #
    analyses_ids = [int(i) for i in ids.split(',')]

    analyses = []
    surface_ct = ContentType.objects.get_for_model(Surface)

    registry = AnalysisRegistry()
    visualization_type = None
    for aid in analyses_ids:
        analysis = Analysis.objects.get(id=aid)

        #
        # Get visualization configuration
        #
        _visualization_app_name, _visualization_type = \
            registry.get_visualization_type_for_function_name(analysis.function.name)
        if visualization_type is None:
            visualization_type = _visualization_type
        else:
            if _visualization_type != visualization_type:
                return HttpResponseNotFound('Cannot combine results of selected analyses into a single download')

        #
        # Check whether user has view permission for requested analysis
        #
        if not analysis.is_visible_for_user(request.user):
            return HttpResponseForbidden()

        #
        # Exclude analysis for surfaces having only one topography
        # (this is useful for averages - the only surface analyses so far -
        # and may be controlled by other means later, if needed)
        #
        if (analysis.subject_type == surface_ct) and (analysis.subject.num_topographies() <= 1):
            continue

        analyses.append(analysis)

    #
    # Dispatch
    #
    spec = 'results'  # could be used to download different things
    key = (visualization_type, spec, file_format)
    try:
        download_function = AnalysisRegistry().get_download_function(*key)
    except UnknownKeyException:
        return HttpResponseBadRequest(
            f"Cannot provide a download for '{spec}' as analysis result type '{visualization_type}' in file format '{file_format}'")

    return download_function(request, analyses)


def analyses_meta_data_dataframe(analyses, request):
    """Generates a pandas.DataFrame with meta data about analyses.

    Parameters
    ----------
    analyses:
        sequence of Analysis instances
    request:
        current request

    Returns
    -------
    pandas.DataFrame, can be inserted as extra sheet
    """

    # Collect DOIs, if any
    dois = sorted(set().union(*[a.dois for a in analyses]))

    properties = []
    values = []
    for i, analysis in enumerate(analyses):

        surfaces = analysis.related_surfaces()
        pubs = [surface.publication for surface in surfaces if surface.is_published]

        if i == 0:
            # list function name and a blank line
            properties = ["Function", ""]
            values = [str(analysis.function), ""]

            if len(dois) > 0:
                properties += ['PLEASE CITE THESE DOIs', '']
                values += [', '.join(dois), '']

        properties += ['Subject type', 'Subject name', 'Creator', 'Instrument name', 'Instrument type',
                       'Instrument parameters', 'Further arguments of analysis function', 'Start time of analysis task',
                       'End time of analysis task', 'Duration of analysis task']

        values += [str(analysis.subject.get_content_type().model), str(analysis.subject.name),
                   str(analysis.subject.creator) if hasattr(analysis.subject, 'creator') else '',
                   str(analysis.subject.instrument_name) if hasattr(analysis.subject, 'instrument_name') else '',
                   str(analysis.subject.instrument_type) if hasattr(analysis.subject, 'instrument_type') else '',
                   str(analysis.subject.instrument_parameters)
                   if hasattr(analysis.subject, 'instrument_parameters') else '',
                   str(analysis.kwargs), str(analysis.start_time),
                   str(analysis.end_time), str(analysis.duration)]

        if analysis.configuration is None:
            properties.append("Versions of dependencies")
            values.append("Unknown. Please recalculate this analysis in order to have version information here.")
        else:
            versions_used = analysis.configuration.versions.order_by('dependency__import_name')

            for version in versions_used:
                properties.append(f"Version of '{version.dependency.import_name}'")
                values.append(f"{version.number_as_string()}")

        for pub in pubs:
            # If a surface was published, the URL is inserted
            properties.append("Publication URL (surface data)")
            values.append(request.build_absolute_uri(pub.get_absolute_url()))

        # We want an empty line on the properties sheet in order to distinguish the topographies
        properties.append("")
        values.append("")

    df = pd.DataFrame({'Property': properties, 'Value': values})

    return df


def publications_urls(request, analyses):
    """Return set of publication URLS for given analyses.

    Parameters
    ----------
    request
        HTTPRequest
    analyses
        seq of Analysis instances
    Returns
    -------
    Set of absolute URLs (strings)
    """
    # Collect publication links, if any
    publication_urls = set()
    related_surfaces = set()
    for a in analyses:
        for surface in a.related_surfaces():
            related_surfaces.add(surface)
    for surface in related_surfaces:
        if surface.is_published:
            pub = surface.publication
            pub_url = request.build_absolute_uri(pub.get_absolute_url())
            publication_urls.add(pub_url)
    return publication_urls


def analysis_header_for_txt_file(analysis, as_comment=True, dois=False):
    """

    Parameters
    ----------
    analysis: Analysis
        Analysis instance for which a header should be generated.
    as_comment: bool
        If True, add '# ' in front of each line

    Returns
    -------
    str
    """

    subject = analysis.subject
    subject_type_str = analysis.subject_type.model.title()
    headline = f"{subject_type_str}: {subject.name}"

    s = f'{headline}\n' + '=' * len(headline) + '\n'
    if hasattr(subject, 'creator'):
        s += f'Creator: {subject.creator}\n'
    if hasattr(subject, 'instrument_name'):
        s += f'Instrument name: {subject.instrument_name}\n'
    if hasattr(subject, 'instrument_type'):
        s += f'Instrument type: {subject.instrument_type}\n'
    if hasattr(subject, 'instrument_parameters'):
        s += f'Instrument parameters: {subject.instrument_parameters}\n'
    s += f'Further arguments of analysis function: {analysis.kwargs}\n' + \
         f'Start time of analysis task: {analysis.start_time}\n' + \
         f'End time of analysis task: {analysis.end_time}\n' + \
         f'Duration of analysis task: {analysis.duration}\n'
    if analysis.configuration is None:
        s += 'Versions of dependencies (like "SurfaceTopography") are unknown for this analysis.\n'
        s += 'Please recalculate in order to have version information here.\n'
    else:
        versions_used = analysis.configuration.versions.order_by('dependency__import_name')

        for version in versions_used:
            s += f"Version of '{version.dependency.import_name}': {version.number_as_string()}\n"
    s += '\n'

    # Write DOIs
    if dois and len(analysis.dois) > 0:
        s += 'IF YOU USE THIS DATA IN A PUBLICATION, PLEASE CITE THE FOLLOWING PAPERS:\n'
        for doi in analysis.dois:
            s += f'- {doi}\n'
        s += '\n'

    if as_comment:
        s = textwrap.indent(s, '# ', predicate=lambda s: True)  # prepend to all lines, also empty

    return s


def _get_si_unit_conversion(result):
    """Return SI units and conversion factors"""

    # Unit conversion tool
    ureg = pint.UnitRegistry()
    ureg.default_format = '~P'  # short and pretty

    # Get original units from result dictionary
    xunit = result['xunit']
    yunit = result['yunit']

    # Convert units to SI
    xconv = 1
    if xunit is not None:
        u = ureg(xunit).to_base_units()
        xunit = str(u.u)
        xconv = u.m
    yconv = 1
    if yunit is not None:
        u = ureg(yunit).to_base_units()
        yunit = str(u.u)
        yconv = u.m
    return xunit, xconv, yunit, yconv


@register_download_function(VIZ_SERIES, 'results', 'txt')
def download_plot_analyses_to_txt(request, analyses):
    """Download plot data for given analyses as CSV file.

        Parameters
        ----------
        request
            HTTPRequest
        analyses
            Sequence of Analysis instances

        Returns
        -------
        HTTPResponse
    """
    # Collect publication links, if any
    publication_urls = publications_urls(request, analyses)

    # Collect DOIs, if any
    dois = sorted(set().union(*[a.dois for a in analyses]))

    # Pack analysis results into a single text file.
    f = io.StringIO()
    for i, analysis in enumerate(analyses):
        if i == 0:
            # Write header
            f.write('# {}\n'.format(analysis.function) +
                    '# {}\n'.format('=' * len(str(analysis.function))))

            # Write DOIs
            if len(dois) > 0:
                f.write('# IF YOU USE THIS DATA IN A PUBLICATION, PLEASE CITE THE FOLLOWING PAPERS:\n')
                for doi in dois:
                    f.write(f"# - {doi}\n")
                f.write("#\n")

            # Write publications
            if len(publication_urls) > 0:
                f.write('#\n')
                f.write('# For these analyses, published data was used. Please visit these URLs for details:\n')
                for pub_url in publication_urls:
                    f.write(f'# - {pub_url}\n')
                f.write('#\n')

        f.write(analysis_header_for_txt_file(analysis))

        result = analysis.result
        xunit, xconv, yunit, yconv = _get_si_unit_conversion(result)

        xunit_str = '' if xunit is None else ' ({})'.format(xunit)
        yunit_str = '' if yunit is None else ' ({})'.format(yunit)
        header = 'Columns: {}{}, {}{}'.format(result['xlabel'], xunit_str, result['ylabel'], yunit_str)

        std_err_y_in_series = any('std_err_y' in s.keys() for s in result['series'])
        if std_err_y_in_series:
            header += ', standard error of average {}{}'.format(result['ylabel'], yunit_str)
            f.writelines([
                '# The value "nan" for the standard error of an average indicates that no error\n',
                '# could be computed because the average contains only a single data point.\n\n'])

        for series in result['series']:
            series_data = [np.array(series['x']) * xconv, np.array(series['y']) * yconv]
            if std_err_y_in_series:
                try:
                    std_err_y = series['std_err_y'] * yconv
                    if hasattr(std_err_y, 'filled'):
                        std_err_y = std_err_y.filled(np.nan)
                    series_data.append(std_err_y)
                except KeyError:
                    pass
            np.savetxt(f, np.transpose(series_data),
                       header='{}\n{}\n{}'.format(series['name'], '-' * len(series['name']), header))
            f.write('\n')

    # Prepare response object.
    response = HttpResponse(f.getvalue(), content_type='application/text')
    filename = f'{slugify(analysis.function.name)}.txt'
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)

    # Close file and return response.
    f.close()
    return response


def _comment_on_average(y, std_err_y_masked):
    """Calculate a comment.

    Parameters:
        y: float
        std_err_y_masked: boolean
    """
    if np.isnan(y):
        return 'average could not be computed'
    elif std_err_y_masked:
        return 'no error could be computed because the average contains only a single data point'
    return ''


@register_download_function(VIZ_SERIES, 'results', 'xlsx')
def download_plot_analyses_to_xlsx(request, analyses):
    """Download plot data for given analyses as XLSX file.

    Parameters
    ----------
    request
        HTTPRequest
    analyses
        Sequence of Analysis instances

    Returns
    -------
    HTTPResponse
    """
    # Pack analysis results into a single text file.
    excel_file_buffer = io.BytesIO()
    excel_writer = pd.ExcelWriter(excel_file_buffer)
    bold_font = Font(bold=True)

    #
    # Create sheet with meta data
    #
    meta_df = analyses_meta_data_dataframe(analyses, request)
    meta_df.to_excel(excel_writer, sheet_name='INFORMATION', index=False, freeze_panes=(1, 0))

    # Analyze subject names and store a distinct name
    # which can be used in sheet names if subject names are not unique
    subject_names_in_sheet_names = [a.subject.name for a in analyses]

    for sn in set(subject_names_in_sheet_names):  # iterate over distinct names

        # replace name with a unique one using a counter
        indices = [i for i, a in enumerate(analyses) if a.subject.name == sn]

        if len(indices) > 1:  # only rename if not unique
            for k, idx in enumerate(indices):
                subject_names_in_sheet_names[idx] += f" ({k + 1})"

    index_entries = []  # tuples with (subject name, subject type, function name, data series, hyperlink to sheet)

    for analysis_idx, analysis in enumerate(analyses):
        result = analysis.result
        xunit, xconv, yunit, yconv = _get_si_unit_conversion(result)
        column1 = '{} ({})'.format(result['xlabel'], xunit)
        column2 = '{} ({})'.format(result['ylabel'], yunit)
        column3 = 'standard error of {} ({})'.format(result['ylabel'], yunit)
        column4 = 'comment'

        creator = str(analysis.subject.creator) if hasattr(analysis.subject, 'creator') else ''
        instrument_name = str(analysis.subject.instrument_name) if hasattr(analysis.subject, 'instrument_name') else ''
        instrument_type = str(analysis.subject.instrument_type) if hasattr(analysis.subject, 'instrument_type') else ''
        instrument_parameters = str(analysis.subject.instrument_parameters) \
            if hasattr(analysis.subject, 'instrument_parameters') else ''

        for series_idx, series in enumerate(result['series']):
            df_columns_dict = {column1: np.array(series['x']) * xconv, column2: np.array(series['y']) * yconv}
            try:
                std_err_y_mask = series['std_err_y'].mask
            except (AttributeError, KeyError) as exc:
                std_err_y_mask = np.zeros(len(series['y']), dtype=bool)

            try:
                df_columns_dict[column3] = np.array(series['std_err_y']) * yconv
                df_columns_dict[column4] = [_comment_on_average(y, masked)
                                            for y, masked in zip(series['y'], std_err_y_mask)]
            except KeyError:
                pass
            df = pd.DataFrame(df_columns_dict)

            #
            # Save data for index entry
            #
            sheet_name = f"analysis-{analysis_idx}-series-{series_idx}"

            subject_type = analysis.subject.get_content_type().name  # human-readable name
            if subject_type == 'topography':
                subject_type = 'measurement'  # this is how topographies are denoted in the UI
            index_entries.append((analysis.subject.name, subject_type,
                                  analysis.function.name, series['name'], sheet_name,
                                  creator, instrument_name, instrument_type, instrument_parameters))

            #
            # Write data sheet to excel file
            #
            df.to_excel(excel_writer, sheet_name=sheet_name, freeze_panes=(10, 1), startcol=0, startrow=9)
            sheet = excel_writer.sheets[sheet_name]
            sheet["A1"] = "Function name"
            sheet["A1"].font = bold_font
            sheet["B1"] = analysis.function.name
            sheet["A2"] = "Subject"
            sheet["A2"].font = bold_font
            sheet["B2"] = analysis.subject.name
            sheet["A3"] = "Subject type"
            sheet["A3"].font = bold_font
            sheet["B3"] = subject_type
            sheet["A4"] = "Creator"
            sheet["A4"].font = bold_font
            sheet["B4"] = creator
            sheet["A5"] = "Instrument name"
            sheet["A5"].font = bold_font
            sheet["B5"] = instrument_name
            sheet["A6"] = "Instrument type"
            sheet["A6"].font = bold_font
            sheet["B6"] = instrument_type
            sheet["A7"] = "Instrument parameters"
            sheet["A7"].font = bold_font
            sheet["B7"] = instrument_parameters
            sheet["A8"] = "Data series"
            sheet["A8"].font = bold_font
            sheet["B8"] = series['name']
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 25

            # Link "Back to Index"
            sheet["D1"].hyperlink = Hyperlink(ref=f"D1",
                                              location="'INDEX'!A1",
                                              tooltip=f"Click to jump back to INDEX")
            sheet["D1"].value = "Click to jump back to INDEX"
            sheet["D1"].style = "Hyperlink"

    excel_writer.close()

    filename = f'{slugify(analysis.function.name)}.xlsx'
    #
    # Create INDEX sheet with links in Openpyxl
    #
    wb = openpyxl.load_workbook(excel_file_buffer)
    excel_file_buffer.close()

    index_ws = wb.create_sheet("INDEX", 0)

    index_headers = ["Subject name", "Subject type", "Function name", "Data series", "Link", "Creator",
                     "Instrument name", "Instrument type", "Instrument parameters"]
    for col_idx, col_header in enumerate(index_headers):
        header_cell = index_ws.cell(row=1, column=col_idx + 1)
        header_cell.value = col_header
        header_cell.font = bold_font

    def create_index_entry(row, index_entry):
        """Create a row on the index sheet."""
        subject_name, subject_type, function_name, data_series, sheet_name, creator, instrument_name, instrument_type, \
            instrument_parameters = index_entry
        index_ws.cell(row=row, column=1).value = subject_name
        index_ws.cell(row=row, column=2).value = subject_type
        index_ws.cell(row=row, column=3).value = function_name
        index_ws.cell(row=row, column=4).value = data_series
        index_ws.cell(row=row, column=6).value = creator
        index_ws.cell(row=row, column=7).value = instrument_name
        index_ws.cell(row=row, column=8).value = instrument_type
        index_ws.cell(row=row, column=9).value = instrument_parameters
        hyperlink_cell = index_ws.cell(row=row, column=5)
        hyperlink_label = f"Click to jump to sheet '{sheet_name}'"
        hyperlink_location = f"'{sheet_name}'!B2"  # don't use # here before sheet name, does not work
        # Hyperlink class: target keyword seems to be used for external links
        # hyperlink_cell.value = f'=HYPERLINK("{hyperlink_location}", "{hyperlink_label}")'
        # didn't manage to get it working with HYPERLINK function
        hyperlink_cell.value = hyperlink_label
        hyperlink_cell.hyperlink = Hyperlink(ref=f"E{row}",
                                             location=hyperlink_location,
                                             tooltip=f"Click to jump to sheet '{sheet_name}' with data")
        hyperlink_cell.style = "Hyperlink"

    for entry_idx, index_entry in enumerate(index_entries):
        create_index_entry(entry_idx + 2, index_entry)
    # create_index_entry(len(index_entries) + 2, ("META DATA", '', '', '', 'META DATA'))

    # increase column width on index page
    for col, colwidth in [("A", 30), ("B", 15), ("C", 25), ("D", 30), ("E", 45)]:
        index_ws.column_dimensions[col].width = colwidth

    #
    # Save to named temporary file and prepare response
    #
    with tempfile.NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)

        response = HttpResponse(tmp.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)

    return response


@register_download_function(VIZ_SERIES, 'results', 'csv')
def download_plot_analyses_to_csv(request, analyses):
    """Download plot data for given analyses as CSV file.

    Parameters
    ----------
    request
        HTTPRequest
    analyses
        Sequence of Analysis instances

    Returns
    -------
    HTTPResponse
    """
    analysis = analyses[0]
    result = analysis.result
    xunit, xconv, yunit, yconv = _get_si_unit_conversion(result)
    column1 = '{} ({})'.format(result['xlabel'], xunit)
    column2 = '{} ({})'.format(result['ylabel'], yunit)
    column3 = 'standard error of {} ({})'.format(result['ylabel'], yunit)
    column4 = 'comment'

    column_subject_type = 'Subject type'
    column_subject_name = 'Subject name'
    column_creator = 'Creator'
    column_instrument_name = 'Instrument name'
    column_instrument_type = 'Instrument type'
    column_instrument_parameters = 'Instrument parameters'
    column_function_name = 'Function name'
    column_data_series = 'Data series'

    data = pd.DataFrame(columns=[
        column_subject_type,
        column_subject_name,
        column_creator,
        column_instrument_name,
        column_instrument_type,
        column_instrument_parameters,
        column_function_name,
        column_data_series,
        column1,
        column2,
        column3,
        column4
    ])

    for analysis_idx, analysis in enumerate(analyses):
        # Get results and compute unit conversion factors
        result = analysis.result
        xunit, xconv, yunit, yconv = _get_si_unit_conversion(result)

        # FIXME! Check that columns are actually identical
        # _column1 = '{} ({})'.format(result['xlabel'], xunit)
        # _column2 = '{} ({})'.format(result['ylabel'], yunit)
        # _column3 = 'standard error of {} ({})'.format(result['ylabel'], yunit)
        # _column4 = 'comment'

        # Get metadata
        subject_type = analysis.subject.get_content_type().name  # human-readable name
        if subject_type == 'topography':
            subject_type = 'measurement'  # this is how topographies are denoted in the UI
        creator = str(analysis.subject.creator) if hasattr(analysis.subject, 'creator') else ''
        instrument_name = str(analysis.subject.instrument_name) if hasattr(analysis.subject, 'instrument_name') else ''
        instrument_type = str(analysis.subject.instrument_type) if hasattr(analysis.subject, 'instrument_type') else ''
        instrument_parameters = str(analysis.subject.instrument_parameters) \
            if hasattr(analysis.subject, 'instrument_parameters') else ''

        for series_idx, series in enumerate(result['series']):
            x = np.array(series['x'])
            y = np.array(series['y'])
            df_columns_dict = {
                column_subject_type: len(x) * [subject_type],
                column_subject_name: len(x) * [analysis.subject.name],
                column_creator: len(x) * [creator],
                column_instrument_name: len(x) * [instrument_name],
                column_instrument_type: len(x) * [instrument_type],
                column_instrument_parameters: len(x) * [instrument_parameters],
                column_function_name: len(x) * [analysis.function.name],
                column_data_series: len(x) * [series['name']],
                column1: x * xconv,
                column2: y * yconv
            }
            try:
                std_err_y_mask = series['std_err_y'].mask
            except (AttributeError, KeyError) as exc:
                std_err_y_mask = np.zeros(len(series['y']), dtype=bool)

            try:
                df_columns_dict[column3] = np.array(series['std_err_y']) * yconv
                df_columns_dict[column4] = [_comment_on_average(y, masked)
                                            for y, masked in zip(series['y'], std_err_y_mask)]
            except KeyError:
                pass

            data = pd.concat([data, pd.DataFrame(df_columns_dict)])

    f = io.StringIO()
    data.to_csv(f, sep=';', index=False)
    response = HttpResponse(f.getvalue(), content_type='application/text')
    filename = f'{slugify(analysis.function.name)}.csv'
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)

    return response
